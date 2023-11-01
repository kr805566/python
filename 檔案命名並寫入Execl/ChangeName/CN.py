import os
from openpyxl import Workbook

def mkdir(path):
    #判斷目錄是否存在
    folder = os.path.exists(path)

    #判斷結果
    if not folder:
        #如果不存在，則建立新目錄
        os.makedirs(path)



if __name__ == '__main__':


    #要改更名的資料放在./before

    beforePath = './before/'
    afterPath = './after/'


    wbTrain = Workbook()
    wbVal = Workbook()
    wbTest = Workbook()

    wbaTrain = wbTrain.active
    wbaVal = wbVal.active
    wbaTest = wbTest.active

    fileList=os.listdir(beforePath)

    sumN = [0, 0, 0]
    sumN2 =[1, 1, 1]

    for i in fileList:
        ImageList = os.listdir(beforePath + i)

        mkdir(afterPath + i)
        mkdir(afterPath + i + '/train')
        mkdir(afterPath + i + '/val')
        mkdir(afterPath + i + '/test')

        print('處理 {} 景點中...'.format(i))
        count = 0
        for j in ImageList:
            count += 1
            if (count <=1050):
                sumN[0] += 1
                str1 = str(count)
                str2 = str1.zfill(4)
                fileName = i
                os.rename(beforePath + i + '/' + j, afterPath + i + '/train/' + fileName + '_train_' + str2 + '.jpg')
                wbaTrain.append(['',fileName + '_train_' + str2 + '.jpg'])
            elif (count <= 1350):
                sumN[1] += 1
                str1 = str(count-1050)
                str2 = str1.zfill(3)
                fileName = i
                os.rename(beforePath + i + '/' + j, afterPath + i + '/val/' + fileName + '_val_' + str2 + '.jpg')
                wbaVal.append(['', fileName + '_val_' + str2 + '.jpg'])
            else:
                sumN[2] += 1
                str1 = str(count-1350)
                str2 = str1.zfill(3)
                fileName = i
                os.rename(beforePath + i + '/' + j, afterPath + i + '/test/' + fileName + '_test_' + str2 + '.jpg')
                wbaTest.append(['', fileName + '_val_' + str2 + '.jpg'])

        #寫入Execl
        wbaTrain['A' + str(sumN2[0])] = fileName
        wbaVal['A' + str(sumN2[1])] = fileName
        wbaTest['A' + str(sumN2[2])] = fileName
        sumN2[0] = 1 + sumN[0]
        sumN2[1] = 1 + sumN[1]
        sumN2[2] = 1 + sumN[2]
    wbTrain.save('Train.xlsx')
    wbVal.save('Val.xlsx')
    wbTest.save('Test.xlsx')
